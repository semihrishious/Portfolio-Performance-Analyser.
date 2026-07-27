import os

from openpyxl import Workbook
from openpyxl.styles import (Font,PatternFill,Alignment,Border,Side)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def create_summary(worksheet,portfolio_return,portfolio_risk,tickers,sharpe_ratio,cagr,
beta,alpha):
    worksheet["A1"] = "Portfolio Report"

    worksheet["A3"] = "Metric"
    worksheet["B3"] = "Value"

    worksheet["A4"] = "Portfolio Return"
    worksheet["B4"] = portfolio_return

    worksheet["A5"] = "Portfolio Risk"
    worksheet["B5"] = portfolio_risk

    worksheet["A6"] = "Portfolio Sharpe Ratio"
    worksheet["B6"] = sharpe_ratio

    worksheet["A7"] = "Portfolio CAGR"
    worksheet["B7"] = cagr

    worksheet["A8"] = "Portfolio Beta"
    worksheet["B8"] = beta

    worksheet["A9"] = "Portfolio Alpha"
    worksheet["B9"] = alpha

    worksheet["A10"] = "Number of Stocks"
    worksheet["B10"] = len(tickers)

 

def generate_excel_report(portfolio_return,portfolio_risk,tickers,
sharpe_ratio,cagr,beta,alpha):
    workbook=Workbook()
    worksheet=workbook.active
    worksheet.title="Portfolio Summary"
    title_font=Font(size=18,bold=True)
    header_font=Font(bold=True)
    header_fill=PatternFill(fill_type="solid",start_color="4F81BD")
    header_font=Font(bold=True,color="FFFFFF")
    center=Alignment(horizontal="center")
    thin=Side(style="thin")
    border=Border(left=thin,right=thin,top=thin,bottom=thin)
    
    create_summary(worksheet,portfolio_return,portfolio_risk,tickers,sharpe_ratio,cagr,beta,alpha)
 
    worksheet["A1"].font=title_font
    worksheet["A3"].font = header_font
    worksheet["A3"].fill = header_fill
    worksheet["A3"].alignment = center
    worksheet["A3"].border = border
    worksheet["B3"].font = header_font
    worksheet["B3"].fill = header_fill
    worksheet["B3"].alignment = center
    worksheet["B3"].border = border
    worksheet["A4"].border = border
    worksheet["B4"].border = border
    worksheet["B4"].number_format = "0.00%"
    worksheet["A5"].border = border
    worksheet["B5"].border = border
    worksheet["B5"].number_format = "0.00%"
    worksheet["A6"].border = border
    worksheet["B6"].border = border
    worksheet["A7"].border = border
    worksheet["B7"].border = border
    worksheet["B7"].number_format = "0.00%"
    worksheet["A8"].border = border
    worksheet["B8"].border = border
    worksheet["A9"].border = border
    worksheet["B9"].border = border
    worksheet["B9"].number_format = "0.00%"
    worksheet["A10"].border = border
    worksheet["B10"].border = border

    chart_path="charts/portfoliovsbenchmark.png"
    if os.path.exists(chart_path):
        chart_image=Image(chart_path)
        chart_image.width = 700
        chart_image.height = 400
        worksheet.add_image(chart_image,"G12")


    for column in worksheet.columns:
        max_length=0
        for cell in column:
            length=len(str(cell.value))
            if length>max_length:
                max_length=length
        column_letter=get_column_letter(column[0].column)
        worksheet.column_dimensions[column_letter].width=max_length+2

    os.makedirs("reports",exist_ok=True)

    workbook.save("reports/Portfolio_Report.xlsx")


